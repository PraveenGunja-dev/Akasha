import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(r'D:\Akasha_Platform\Data\NEW31\ZPSPS007 (3).xlsx', read_only=True)
ws = wb.active

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

# Collect unique values for key columns
types = Counter()       # col 4 = Type
doc_types = Counter()   # col 9 = Document Type
wbs_prefixes = Counter()  # first part of WBS Element
summaries = Counter()     # col 0 = Summary

row_count = 0
sample_rows_by_type = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    row_count += 1
    t = str(row[4] or '').strip()
    dt = str(row[9] or '').strip()
    wbs = str(row[1] or '').strip()
    summary = str(row[0] or '').strip()
    
    types[t] += 1
    doc_types[dt] += 1
    summaries[summary] += 1
    
    # WBS prefix (e.g. H-6061, H-5120)
    if wbs:
        parts = wbs.split('-')
        if len(parts) >= 2:
            wbs_prefixes[f'{parts[0]}-{parts[1]}'] += 1
    
    # Collect a sample row per type
    if t not in sample_rows_by_type:
        sample_rows_by_type[t] = list(row)

print(f'Total data rows: {row_count}')

print(f'\n--- Type (col 4) ---')
for k, v in types.most_common(20):
    print(f'  "{k}": {v}')

print(f'\n--- Document Type (col 9) ---')
for k, v in doc_types.most_common(20):
    print(f'  "{k}": {v}')

print(f'\n--- WBS Prefixes (top 20) ---')
for k, v in wbs_prefixes.most_common(20):
    print(f'  "{k}": {v}')

print(f'\n--- Summary (col 0) ---')
for k, v in summaries.most_common(10):
    print(f'  "{k}": {v}')

print(f'\n--- Sample rows by Type ---')
for t, row in sample_rows_by_type.items():
    print(f'\n  Type="{t}":')
    for i, h in enumerate(headers):
        print(f'    {h}: {row[i]}')

wb.close()
