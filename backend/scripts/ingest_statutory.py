import os
import sys
import openpyxl
from datetime import datetime

# Add the parent directory (backend) to the Python path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from database import SessionLocal
import models

DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../Data/NEW31"))
SAP_MASTER_FILE = os.path.join(DATA_DIR, "AKASHA SAP MASTER FILE (2).xlsx")
STATUTORY_STATUS_FILE = os.path.join(DATA_DIR, "Statutory Status Khavda.xlsx")
EPC_BOCW_FILE = os.path.join(DATA_DIR, "EPC Partners- BOCW, CLRA & GST Status.xlsx")
INSURANCE_MASTER_FILE = os.path.join(DATA_DIR, "Insurance Master List Phase IV.xlsx")

def load_sap_mapping():
    """Loads mapping from SPV+Plot to project_id and p6_project_name"""
    wb = openpyxl.load_workbook(SAP_MASTER_FILE, data_only=True)
    ws = wb.active
    sap_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        p6_id = str(row[0]).strip() if row[0] else ''
        proj_name = str(row[1]).strip() if row[1] else ''
        spv = str(row[2]).strip() if row[2] else ''
        if p6_id and spv:
            sap_map[(spv, proj_name)] = (p6_id, proj_name)
    return sap_map

def find_mapping(spv, plot, sap_map):
    """Finds project_id and p6_project_name based on SPV and Plot string matching"""
    if not spv:
        return None, None
        
    plot_upper = plot.upper() if plot else ""
    
    for (s, pn), (pid, pname) in sap_map.items():
        pn_upper = pn.upper()
        if s == spv and plot_upper in pn_upper:
            return pid, pname
        # Fallback exact plot wrapper matching
        if s == spv and f"_{plot_upper}_" in f"_{pn_upper}_":
            return pid, pname
            
    return None, None

def safe_float(val):
    try:
        return float(val) if val else None
    except:
        return None

def safe_date(val):
    if isinstance(val, datetime):
        return val
    return None

def ingest_statutory_compliance(db, sap_map):
    print("Ingesting Statutory Compliance...")
    wb = openpyxl.load_workbook(STATUTORY_STATUS_FILE, data_only=True)
    ws = wb['Statutory']
    
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        proj = str(row[0]).strip() if row[0] else None
        if not proj: continue
        
        spv = str(row[1]).strip() if row[1] else None
        plot = str(row[2]).strip() if row[2] else None
        cat = str(row[3]).strip() if row[3] else None
        type_str = str(row[4]).strip() if row[4] else None
        vendor = str(row[5]).strip() if row[5] else None
        ol = safe_float(row[6])
        cap = safe_float(row[7])
        gst = str(row[8]).strip() if row[8] else None
        bocw = str(row[9]).strip() if row[9] else None
        clra = str(row[10]).strip() if row[10] else None
        spcb = str(row[11]).strip() if row[11] else None
        sublease = str(row[12]).strip() if row[12] else None
        ins = str(row[13]).strip() if row[13] else None
        
        project_id, p6_project_name = find_mapping(spv, plot, sap_map)
        
        record = models.StatutoryCompliance(
            project_id=project_id,
            p6_project_name=p6_project_name,
            project_name=proj,
            spv_code=spv,
            plot_location=plot,
            epc_partner=vendor,
            category=cat,
            module_type=type_str,
            capacity_mwac=cap,
            ol=ol,
            gst_status=gst,
            bocw_status=bocw,
            clra_status=clra,
            spcb_status=spcb,
            sub_lease_status=sublease,
            insurance_status=ins,
            upload_time=datetime.utcnow()
        )
        db.add(record)
        count += 1
    
    db.commit()
    print(f"Added {count} records to StatutoryCompliance.")

def ingest_epc_bocw(db, sap_map):
    print("Ingesting EPC BOCW & CLRA...")
    wb = openpyxl.load_workbook(EPC_BOCW_FILE, data_only=True)
    ws = wb['BOCW']
    
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        epc = str(row[1]).strip() if row[1] else None
        if not epc or epc == 'None': continue
        
        proj = str(row[2]).strip() if row[2] else None
        spv = str(row[4]).strip() if row[4] else None
        plot = str(row[5]).strip() if row[5] else None
        cap = safe_float(row[6])
        ftc = safe_date(row[7])
        bocw_ent = str(row[8]).strip() if row[8] else None
        bocw_comm = safe_date(row[9])
        bocw_val = safe_date(row[10])
        
        project_id, p6_project_name = find_mapping(spv, plot, sap_map)
        
        # We assume CLRA and GST will be updated or joined.
        # This is a simplified ingestion for demo purposes.
        record = models.EPCStatutoryStatus(
            project_id=project_id,
            p6_project_name=p6_project_name,
            epc_partner=epc,
            project_name=proj,
            spv_code=spv,
            plot=plot,
            capacity_mw=cap,
            ftc_date=ftc,
            bocw_license_entity=bocw_ent,
            bocw_commencement_date=bocw_comm,
            bocw_validity_date=bocw_val,
            upload_time=datetime.utcnow()
        )
        db.add(record)
        count += 1
        
    db.commit()
    print(f"Added {count} records to EPCStatutoryStatus.")

def ingest_insurance(db, sap_map):
    print("Ingesting Insurance Data...")
    wb = openpyxl.load_workbook(INSURANCE_MASTER_FILE, data_only=True)
    ws = wb['Master Sheet']
    
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        proj = str(row[1]).strip() if row[1] else None
        if not proj or proj == 'None': continue
        
        spv = str(row[2]).strip() if row[2] else None
        plot = str(row[3]).strip() if row[3] else None
        cat = str(row[4]).strip() if row[4] else None
        m_type = str(row[5]).strip() if row[5] else None
        vendor = str(row[6]).strip() if row[6] else None
        cap_ac = safe_float(row[7])
        cap_dc = safe_float(row[8])
        ins_co = str(row[9]).strip() if row[9] else None
        sum_ins = safe_float(row[10])
        cov_let = str(row[11]).strip().lower() == 'yes' if row[11] else False
        pol_no = str(row[12]).strip() if row[12] else None
        pol_start = safe_date(row[13])
        pol_exp = safe_date(row[14])
        prem = safe_float(row[15])
        rem = str(row[16]).strip() if row[16] else None
        alert = str(row[17]).strip() if row[17] else None
        
        project_id, p6_project_name = find_mapping(spv, plot, sap_map)
        
        record = models.InsurancePolicy(
            project_id=project_id,
            p6_project_name=p6_project_name,
            project_name=proj,
            spv_code=spv,
            plot_location=plot,
            category=cat,
            module_type=m_type,
            epc_vendor=vendor,
            capacity_mwac=cap_ac,
            capacity_mwdc=cap_dc,
            insurance_company=ins_co,
            sum_insured=sum_ins,
            covering_letter=cov_let,
            policy_number=pol_no,
            policy_start=pol_start,
            policy_expiry=pol_exp,
            premium_incl_gst=prem,
            remarks=rem,
            renewal_alert=alert,
            upload_time=datetime.utcnow()
        )
        db.add(record)
        count += 1
        
    db.commit()
    print(f"Added {count} records to InsurancePolicy.")

def main():
    db = SessionLocal()
    
    try:
        print("Clearing old records...")
        db.query(models.StatutoryCompliance).delete()
        db.query(models.EPCStatutoryStatus).delete()
        db.query(models.InsurancePolicy).delete()
        db.commit()
        
        print("Loading SAP Mapping...")
        sap_map = load_sap_mapping()
        
        ingest_statutory_compliance(db, sap_map)
        ingest_epc_bocw(db, sap_map)
        ingest_insurance(db, sap_map)
        
        print("Ingestion completed successfully!")
    except Exception as e:
        print(f"Error: {e}")
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    main()
