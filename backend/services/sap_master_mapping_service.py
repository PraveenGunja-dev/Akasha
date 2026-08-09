"""Validated import of project-to-SAP rules from the Akasha SAP master."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

import models


MASTER_HEADERS = (
    "P6 ID",
    "Project Name",
    "SPV",
    "Type (Cluster)",
    "SPV",
    "AGEL",
    "AGE6L",
    "Plant code ",
    "SPV",
    "AGEL",
    "AGE6L",
)
WBS_PATTERN = re.compile(r"GWL64\.\d+|H-[A-Z0-9]+(?:-[A-Z0-9]+)*", re.IGNORECASE)
EMPTY_MARKERS = {"", "-", "nan", "none", "null", "not found"}
NAME_OWNER_OVERRIDES = {"ASEJ6PL_S07_FT_300MW_PPA": "FY26-P14"}


@dataclass(frozen=True)
class SapMasterSyncReport:
    projects_seen: int
    projects_inserted: int
    projects_updated: int
    scopes_written: int
    shared_scope_groups: int
    intentionally_unmapped: int
    warnings: tuple[str, ...]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    cleaned = str(value).strip()
    return "" if cleaned.casefold() in EMPTY_MARKERS else cleaned


def _project_id(value: Any) -> str:
    return re.sub(r"\s+", "", _text(value)).upper()


def _number(value: Any) -> float | None:
    try:
        return float(value) if _text(value) else None
    except (TypeError, ValueError):
        return None


def _wbs_roots(value: Any) -> tuple[str, ...]:
    roots = tuple(dict.fromkeys(match.upper() for match in WBS_PATTERN.findall(_text(value))))
    if _text(value) and not roots:
        raise ValueError(f"Invalid WBS mapping value: {_text(value)!r}")
    return roots


def _plant_code(value: Any) -> str | None:
    cleaned = _text(value)
    return cleaned.upper() or None


def _read_rows(path: Path) -> tuple[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = tuple(sheet.cell(1, column).value for column in range(1, 12))
    if headers != MASTER_HEADERS:
        raise ValueError(f"Unexpected SAP master columns: {headers!r}")
    rows = []
    for row_number in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, 12)]
        project_id = _project_id(values[0])
        if not project_id:
            continue
        rows.append({
            "source_row": row_number,
            "project_id": project_id,
            "project_name": _text(values[1]),
            "spv_name": _text(values[2]),
            "cluster": _text(values[3]),
            "wbs": {
                "SPV": _wbs_roots(values[4]),
                "AGEL": _wbs_roots(values[5]),
                "AGE6L": _wbs_roots(values[6]),
            },
            "primary_plant": _plant_code(values[7]),
            "wind_plants": {
                "SPV": _plant_code(values[8]),
                "AGEL": _plant_code(values[9]),
                "AGE6L": _plant_code(values[10]),
            },
        })
    workbook.close()
    return sheet.title, rows


def _read_legacy(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        project_id = _project_id(row.get("Project ID"))
        if project_id:
            rows[project_id] = row
    workbook.close()
    return rows


def _legacy_text(row: dict[str, Any], key: str) -> str | None:
    return _text(row.get(key)) or None


def _unique_project_names(rows: list[dict[str, Any]]) -> list[str]:
    owners: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        if row["project_name"]:
            owners[row["project_name"]].append(row["project_id"])
    warnings = []
    for name, project_ids in owners.items():
        if len(project_ids) == 1:
            continue
        owner = NAME_OWNER_OVERRIDES.get(name)
        if owner not in project_ids:
            raise ValueError(f"Duplicate project name {name!r}: {project_ids}")
        for row in rows:
            if row["project_name"] == name and row["project_id"] != owner:
                row["project_name"] = row["project_id"]
        warnings.append(f"{name} assigned to {owner}; other duplicate rows use their P6 ID.")
    return warnings


def sync_sap_master(
    db: Session,
    master_path: str | Path,
    legacy_path: str | Path,
    *,
    dry_run: bool = False,
) -> SapMasterSyncReport:
    master_path = Path(master_path)
    legacy_path = Path(legacy_path)
    if not master_path.exists():
        raise FileNotFoundError(master_path)

    sheet_name, rows = _read_rows(master_path)
    warnings = _unique_project_names(rows)
    legacy = _read_legacy(legacy_path)
    existing_rows = db.query(models.ProjectMapping).order_by(models.ProjectMapping.id).all()
    existing: dict[str, models.ProjectMapping] = {}
    for mapping in existing_rows:
        key = _project_id(mapping.project_id)
        if key in existing:
            raise ValueError(f"Duplicate project_mapping project_id after normalization: {key}")
        existing[key] = mapping

    inserted = 0
    updated = 0
    mappings: dict[str, models.ProjectMapping] = {}
    for row in rows:
        project_id = row["project_id"]
        old = legacy.get(project_id, {})
        mapping = existing.get(project_id)
        if mapping is None:
            mapping = models.ProjectMapping(project_id=project_id)
            db.add(mapping)
            inserted += 1
        else:
            updated += 1

        mapping.project_id = project_id
        mapping.project_name_from_p6 = row["project_name"] or project_id
        mapping.project = _legacy_text(old, "Project") or mapping.project or mapping.project_name_from_p6
        mapping.spv_name = row["spv_name"] or mapping.spv_name
        mapping.cluster = row["cluster"] or mapping.cluster
        mapping.plot_no = _legacy_text(old, "Plot No") or mapping.plot_no
        mapping.category = _legacy_text(old, "Category") or mapping.category
        mapping.mms_type = _legacy_text(old, "MMS Type") or mapping.mms_type
        mapping.ol = _legacy_text(old, "OL") or mapping.ol
        mapping.not_allocated = _legacy_text(old, "Not Allocated") or mapping.not_allocated
        mapping.priority = _legacy_text(old, "Priority") or mapping.priority
        mapping.source_of_origin = _legacy_text(old, "SourceOfOrigin") or mapping.source_of_origin
        mapping.capacity_mwac = _number(old.get("Capacity\n(MWac)")) or mapping.capacity_mwac
        mapping.capacity_mwdc = _number(old.get("Capacity (MWdc)")) or mapping.capacity_mwdc
        mapping.module_wbs = _legacy_text(old, "Module WBS") or mapping.module_wbs

        wind_plants = row["wind_plants"] if "wind" in row["cluster"].casefold() else {}
        mapping.spv_plant_code = (
            row["primary_plant"]
            or wind_plants.get("SPV")
            or _legacy_text(old, "SPVPlantCode -  machinery")
            or mapping.spv_plant_code
        )
        mapping.agel = (
            wind_plants.get("AGEL")
            or _legacy_text(old, "AGEL - Supplay material")
            or mapping.agel
        )
        mapping.age6l = (
            wind_plants.get("AGE6L")
            or _legacy_text(old, "AGE6L -")
            or _legacy_text(old, "AGE6L - ")
            or mapping.age6l
        )
        mappings[project_id] = mapping

    db.flush()
    mapping_ids = [mapping.id for mapping in mappings.values()]
    if mapping_ids:
        db.query(models.SapProjectScope).filter(
            models.SapProjectScope.project_mapping_id.in_(mapping_ids)
        ).delete(synchronize_session=False)

    pending_rules = []
    for row in rows:
        mapping = mappings[row["project_id"]]
        for owner, roots in row["wbs"].items():
            for root in roots:
                pending_rules.append((mapping, owner, "wbs_prefix", root, row["source_row"]))
        if "wind" in row["cluster"].casefold():
            for owner, plant in row["wind_plants"].items():
                if plant:
                    pending_rules.append((mapping, owner, "plant_code", plant, row["source_row"]))

    groups: dict[tuple[str, str, str], list[tuple]] = defaultdict(list)
    for rule in pending_rules:
        groups[(rule[1], rule[2], rule[3])].append(rule)

    shared_groups = 0
    for group, rules in groups.items():
        capacities = [float(rule[0].capacity_mwac or 0) for rule in rules]
        if len(rules) > 1:
            shared_groups += 1
            if any(capacity <= 0 for capacity in capacities):
                project_ids = [rule[0].project_id for rule in rules]
                if group[1] == "wbs_prefix":
                    raise ValueError(f"Shared SAP scope {group} lacks capacity: {project_ids}")
                capacities = [1.0] * len(rules)
                warnings.append(
                    f"Shared wind plant scope {group} used equal allocation because capacity "
                    f"was unavailable for: {project_ids}."
                )
        total_capacity = sum(capacities)
        for rule, capacity in zip(rules, capacities):
            weight = capacity / total_capacity if len(rules) > 1 else 1.0
            db.add(models.SapProjectScope(
                project_mapping_id=rule[0].id,
                owner=rule[1],
                match_kind=rule[2],
                match_value=rule[3],
                allocation_group=":".join(group),
                allocation_weight=weight,
                source_file=master_path.name,
                source_sheet=sheet_name,
                source_row=rule[4],
            ))

    db.flush()
    report = SapMasterSyncReport(
        projects_seen=len(rows),
        projects_inserted=inserted,
        projects_updated=updated,
        scopes_written=len(pending_rules),
        shared_scope_groups=shared_groups,
        intentionally_unmapped=sum(
            not any(row["wbs"].values())
            and not (
                "wind" in row["cluster"].casefold()
                and any(row["wind_plants"].values())
            )
            for row in rows
        ),
        warnings=tuple(warnings),
    )
    if dry_run:
        db.rollback()
    return report
